# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import json

wb = load_workbook(r'C:\Users\iceam\Downloads\payroll_temp.xlsx')

result = {'sheets': []}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_data = {'name': sheet_name, 'rows': []}

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        row_data = {}
        for j, cell in enumerate(row[:30]):
            if cell:
                row_data[f'col_{j+1}'] = str(cell)[:50]
        if row_data:
            sheet_data['rows'].append({'row': i+1, 'data': row_data})

    result['sheets'].append(sheet_data)

with open('payroll_structure.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"Sheets: {wb.sheetnames}")
print("Saved to payroll_structure.json")
