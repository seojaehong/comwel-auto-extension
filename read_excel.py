# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import json

wb = load_workbook(r'C:\Users\iceam\Downloads\근로자고용취득신고_전자신고용 (2).xlsx')
ws = wb.active

result = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
    row_data = {}
    for j, cell in enumerate(row):
        if cell:
            row_data[f'col_{j+1}'] = str(cell)
    if row_data:
        result.append({'row': i+1, 'data': row_data})

# Save as JSON
with open('excel_content.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print("Saved to excel_content.json")
